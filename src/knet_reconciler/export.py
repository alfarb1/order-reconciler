"""Excel + CSV export of the reconciliation result. Per SPEC §6.9.

The xlsx is the thing the user actually looks at — every other piece of the
pipeline exists to populate this report. Optimize for someone scanning it
in Excel: Missing Orders first (it's the actionable list), wide retailer/item
columns, plain dates without microseconds, the noisy debug columns moved out
of the way."""
from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .db import Match, Receipt, Shipment
from .reconcile import FLAG_MISSING, FLAG_PENDING, MATCH_TRACKING, orphan_receipts

GREEN = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
YELLOW = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
RED = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
BANNER_FILL = PatternFill("solid", start_color="FFC000", end_color="FFC000")
HEADER_FILL = PatternFill("solid", start_color="305496", end_color="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F2937")
ZEBRA = PatternFill("solid", start_color="F2F4F8", end_color="F2F4F8")

DATE_FORMAT = "yyyy-mm-dd"

# Each column entry: (key, header, width, number_format)
#   key matches a field on ShipmentRow.as_dict()
ACTIONABLE_COLS = [
    ("retailer",        "Retailer",      24, None),
    ("order_number",    "Order #",       18, None),
    ("ship_date",       "Shipped",       12, DATE_FORMAT),
    ("carrier",         "Carrier",        9, None),
    ("tracking",        "Tracking",      24, None),
    ("item",            "Item",          40, None),
    ("subject",         "Email subject", 40, None),
]
ALL_COLS = [
    ("retailer",        "Retailer",      24, None),
    ("order_number",    "Order #",       18, None),
    ("ship_date",       "Shipped",       12, DATE_FORMAT),
    ("received_at",     "KNET received", 12, DATE_FORMAT),
    ("days_to_receipt", "Days",           6, None),
    ("carrier",         "Carrier",        9, None),
    ("tracking",        "Tracking",      24, None),
    ("status",          "Status",        14, None),
    ("item",            "Item",          36, None),
]
ORPHAN_COLS = [
    ("received_at",     "KNET received", 12, DATE_FORMAT),
    ("carrier",         "Carrier",        9, None),
    ("tracking",        "Tracking",      24, None),
    ("notes",           "Notes",         40, None),
]


@dataclass
class ShipmentRow:
    shipment: Shipment
    match: Match | None
    receipt: Receipt | None
    subject: str | None = None

    def as_dict(self) -> dict:
        s, m, r = self.shipment, self.match, self.receipt
        days = None
        if r and s.ship_date and r.received_at:
            try:
                days = (r.received_at - s.ship_date).days
            except TypeError:
                days = None
        if m and m.match_type == MATCH_TRACKING:
            status = "matched"
        elif m and m.flagged_reason == FLAG_MISSING:
            status = FLAG_MISSING
        elif m and m.flagged_reason == FLAG_PENDING:
            status = FLAG_PENDING
        else:
            status = "—"
        return {
            "retailer": s.retailer or "—",
            "order_number": s.order_number or "—",
            "ship_date": _strip_tz(s.ship_date),
            "received_at": _strip_tz(r.received_at) if r else None,
            "days_to_receipt": days,
            "carrier": (s.carrier or "—").upper(),
            "tracking": s.tracking_number_normalized or "—",
            "status": status,
            "item": s.item_description or "—",
            "subject": self.subject or "—",
        }


def _strip_tz(dt):
    if isinstance(dt, datetime) and dt.tzinfo:
        return dt.replace(tzinfo=None)
    return dt


def _gather(session: Session) -> list[ShipmentRow]:
    rows: list[ShipmentRow] = []
    from .db import Email
    for s in session.query(Shipment).all():
        match = session.query(Match).filter(Match.shipment_id == s.id).one_or_none()
        receipt = None
        if match and match.receipt_id:
            receipt = session.get(Receipt, match.receipt_id)
        email = session.get(Email, s.email_id) if s.email_id else None
        subject = email.subject if email else None
        rows.append(ShipmentRow(s, match, receipt, subject))
    return rows


def _write_header(ws, columns, start_row=1):
    for col_idx, (_, header, width, _fmt) in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[start_row].height = 22
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1).coordinate


def _write_data_rows(ws, columns, rows, start_row=2):
    for r_idx, row_data in enumerate(rows, start=start_row):
        zebra = (r_idx - start_row) % 2 == 1
        for c_idx, (key, _header, _width, fmt) in enumerate(columns, start=1):
            val = row_data.get(key)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if fmt:
                cell.number_format = fmt
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if zebra:
                cell.fill = ZEBRA
    return start_row + len(rows) - 1


def _write_banner(ws, text, columns):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = Font(bold=True, size=12, color="3B1F00")
    cell.fill = BANNER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28


def write_xlsx(session: Session, output: Path, stale_days: int = 14) -> Path:
    rows = _gather(session)
    output.parent.mkdir(parents=True, exist_ok=True)

    matched_rows  = [r for r in rows if r.match and r.match.match_type == MATCH_TRACKING]
    missing_rows  = sorted(
        [r for r in rows if r.match and r.match.flagged_reason == FLAG_MISSING],
        key=lambda r: r.shipment.ship_date or datetime.min,
    )
    pending_rows  = sorted(
        [r for r in rows if r.match and r.match.flagged_reason == FLAG_PENDING],
        key=lambda r: r.shipment.ship_date or datetime.min,
    )
    orphans = orphan_receipts(session)

    wb = Workbook()
    _write_summary(
        wb.active,
        total=len(rows),
        matched=len(matched_rows),
        missing=len(missing_rows),
        pending=len(pending_rows),
        orphans=len(orphans),
        stale_days=stale_days,
    )

    # Missing Orders is the actionable list — first sheet after summary.
    ws = wb.create_sheet("Missing Orders")
    _write_banner(
        ws,
        f"Send this list to support@knetgroup.com — {len(missing_rows)} shipments delivered to KNET with no 'received' confirmation",
        ACTIONABLE_COLS,
    )
    _write_header(ws, ACTIONABLE_COLS, start_row=2)
    last = _write_data_rows(ws, ACTIONABLE_COLS, [r.as_dict() for r in missing_rows], start_row=3)

    ws = wb.create_sheet("Pending")
    _write_banner(ws, f"{len(pending_rows)} shipments shipped recently, not yet confirmed by KNET", ACTIONABLE_COLS)
    _write_header(ws, ACTIONABLE_COLS, start_row=2)
    _write_data_rows(ws, ACTIONABLE_COLS, [r.as_dict() for r in pending_rows], start_row=3)

    ws = wb.create_sheet("All Shipments")
    _write_header(ws, ALL_COLS, start_row=1)
    _write_data_rows(ws, ALL_COLS, [r.as_dict() for r in rows], start_row=2)
    _apply_status_colors(ws, ALL_COLS, n_rows=len(rows), start_row=2)

    ws = wb.create_sheet("Orphan Receipts (KNET-side)")
    _write_banner(
        ws,
        f"{len(orphans)} packages KNET received that we have no outbound shipment record for",
        ORPHAN_COLS,
    )
    _write_header(ws, ORPHAN_COLS, start_row=2)
    orphan_rows = [
        {
            "received_at": _strip_tz(r.received_at),
            "carrier": (r.carrier or "—").upper(),
            "tracking": r.tracking_number_normalized or "—",
            "notes": r.notes or "—",
        }
        for r in orphans
    ]
    _write_data_rows(ws, ORPHAN_COLS, orphan_rows, start_row=3)

    wb.save(output)
    return output


def _write_summary(ws, *, total, matched, missing, pending, orphans, stale_days):
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    ws["A1"] = "KNET Reconciliation"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    ws["A2"] = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:B2")

    rows = [
        ("Total shipments",       total),
        ("Matched",               matched),
        ("Missing (action)",      missing),
        ("Pending (in transit)",  pending),
        ("Orphan KNET receipts",  orphans),
        ("Stale threshold (days)", stale_days),
    ]
    for i, (label, value) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=value)
        c.alignment = Alignment(horizontal="right")
        if label == "Missing (action)" and value:
            c.fill = RED
            c.font = Font(bold=True)
        elif label == "Matched" and value:
            c.fill = GREEN

    border = Border(left=Side(style="thin", color="D1D5DB"),
                    right=Side(style="thin", color="D1D5DB"),
                    top=Side(style="thin", color="D1D5DB"),
                    bottom=Side(style="thin", color="D1D5DB"))
    for row in ws.iter_rows(min_row=4, max_row=4 + len(rows) - 1, min_col=1, max_col=2):
        for cell in row:
            cell.border = border


def _apply_status_colors(ws, columns, *, n_rows, start_row):
    if n_rows == 0:
        return
    keys = [c[0] for c in columns]
    if "status" not in keys:
        return
    col_idx = keys.index("status") + 1
    letter = get_column_letter(col_idx)
    last = start_row + n_rows - 1
    rng = f"{letter}{start_row}:{letter}{last}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"matched"'], fill=GREEN))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{FLAG_MISSING}"'], fill=RED))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{FLAG_PENDING}"'], fill=YELLOW))


def write_csv(session: Session, output: Path) -> Path:
    rows = _gather(session)
    output.parent.mkdir(parents=True, exist_ok=True)
    keys = [c[0] for c in ALL_COLS]
    with output.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow([c[1] for c in ALL_COLS])
        for r in rows:
            d = r.as_dict()
            w.writerow([d.get(k) for k in keys])
    return output
